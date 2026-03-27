import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
from datetime import datetime
from PIL import Image
import warnings
import os

warnings.filterwarnings("ignore")

DATA_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "FARMHUB RECORDS.xlsx")
LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.jpeg")

_logo_img = Image.open(LOGO_PATH)
st.set_page_config(
    page_title="FarmHub Agro Services — Farm Dashboard",
    page_icon=_logo_img,
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
/* ── BASE ───────────────────────────────────────────────────────────────────── */
html, body, [data-testid="stAppViewContainer"] { background-color: #ffffff !important; }
[data-testid="stMain"] { background-color: #ffffff !important; }
.block-container { padding-top: 1.8rem !important; padding-bottom: 2rem !important; }

/* ── SIDEBAR ────────────────────────────────────────────────────────────────── */
section[data-testid="stSidebar"] {
    background: linear-gradient(175deg, #0a2010 0%, #122d1a 40%, #1a4526 80%, #1f5c2e 100%) !important;
    border-right: 3px solid #22b545 !important;
}
[data-testid="stSidebarHeader"] { background: transparent !important; border: none !important; }
[data-testid="stSidebarSearch"] { display: none !important; }
section[data-testid="stSidebar"] * { color: #d4f0da !important; }
section[data-testid="stSidebar"] strong { color: #ffffff !important; }
section[data-testid="stSidebar"] hr { border-color: rgba(34,181,69,0.3) !important; }
section[data-testid="stSidebar"] code {
    background: rgba(34,181,69,0.18) !important;
    color: #aee8bc !important; border-radius: 4px; padding: 1px 6px;
    font-size: 0.82rem; border: 1px solid rgba(34,181,69,0.3);
}
/* Radio nav items */
section[data-testid="stSidebar"] label {
    background: transparent !important;
    border-radius: 8px !important;
    padding: 6px 10px !important;
    margin: 2px 0 !important;
    transition: background 0.15s;
    font-size: 0.93rem !important;
    font-weight: 500 !important;
}
section[data-testid="stSidebar"] label:hover { background: rgba(34,181,69,0.15) !important; }
section[data-testid="stSidebar"] input[type="radio"] { accent-color: #22b545 !important; }

/* ── SIDEBAR COLLAPSED BUTTON (mobile — keep visible) ────────────────────────── */
[data-testid="stSidebarCollapsedControl"] {
    background-color: #22b545 !important;
    border-radius: 0 8px 8px 0 !important;
    min-width: 36px !important;
}
[data-testid="stSidebarCollapsedControl"] svg { fill: #ffffff !important; color: #ffffff !important; }
[data-testid="stSidebarCollapsedControl"] button { color: #ffffff !important; }

/* ── MULTISELECT PILLS ────────────────────────────────────────────────────────── */
[data-baseweb="tag"] {
    background-color: rgba(34,181,69,0.18) !important;
    border: 1px solid #22b545 !important;
    border-radius: 6px !important;
}
[data-baseweb="tag"] span { color: #1a5c2a !important; font-weight: 500 !important; }
[data-baseweb="tag"] [role="button"] svg { fill: #2d7a3a !important; }

/* ── TOPBAR ─────────────────────────────────────────────────────────────────── */
header[data-testid="stHeader"] {
    background: #ffffff !important;
    border-bottom: 2px solid #e8f5e9 !important;
}

/* ── PAGE HEADINGS ──────────────────────────────────────────────────────────── */
h1 { color: #1a5c2a !important; font-weight: 700 !important; letter-spacing: -0.3px; }
h2, h3 { color: #2d7a3a !important; font-weight: 600 !important; }
p em { color: #5c8c65 !important; }

/* ── METRIC CARDS ───────────────────────────────────────────────────────────── */
[data-testid="stMetric"] {
    background: #ffffff !important;
    border: 1px solid #e0f0e5 !important;
    border-top: 4px solid #2d7a3a !important;
    border-radius: 10px !important;
    padding: 14px 16px !important;
    box-shadow: 0 2px 8px rgba(45,122,58,0.08) !important;
}
[data-testid="stMetricLabel"] { color: #5c8c65 !important; font-size: 0.72rem !important; font-weight: 600 !important; text-transform: uppercase; letter-spacing: 0.3px; white-space: normal !important; overflow: visible !important; text-overflow: unset !important; line-height: 1.3 !important; }
[data-testid="stMetricValue"] { color: #1a5c2a !important; font-size: 1.35rem !important; font-weight: 700 !important; white-space: normal !important; word-break: break-word; }
[data-testid="stMetricDelta"] { font-size: 0.78rem !important; }

/* ── SECTION HEADERS ────────────────────────────────────────────────────────── */
.section-header {
    font-size: 0.92rem; font-weight: 700; color: #1a5c2a;
    margin-bottom: 10px; padding: 6px 12px;
    background: #f0f9f2;
    border-left: 4px solid #2d7a3a;
    border-radius: 0 6px 6px 0;
    text-transform: uppercase; letter-spacing: 0.4px;
}

/* ── INSIGHT / EVENT CARDS ──────────────────────────────────────────────────── */
.insight-card {
    background: #ffffff;
    border: 1px solid #d4edda;
    border-left: 4px solid #2d7a3a;
    border-radius: 8px;
    padding: 12px 16px;
    margin-bottom: 8px;
    font-size: 0.87rem;
    color: #1a2e1c;
    box-shadow: 0 1px 4px rgba(45,122,58,0.06);
    line-height: 1.5;
}
.insight-card b { color: #1a5c2a; }

/* Upcoming events — amber left border */
.insight-card.upcoming { border-left-color: #e67e22 !important; border-color: #fde8d0 !important; }

/* ── ANIMAL PROFILE CARDS ───────────────────────────────────────────────────── */
.animal-card {
    background: #ffffff;
    border: 1px solid #e0f0e5;
    border-radius: 10px;
    padding: 14px 16px;
    margin-bottom: 10px;
    font-size: 0.83rem;
    color: #1a2e1c;
    box-shadow: 0 2px 6px rgba(45,122,58,0.07);
    line-height: 1.6;
}
.animal-card h4 {
    color: #1a5c2a; margin: 0 0 8px 0; font-size: 1rem; font-weight: 700;
    padding-bottom: 6px; border-bottom: 1px solid #e8f5e9;
}

/* ── SIDEBAR LOGO ────────────────────────────────────────────────────────────── */
section[data-testid="stSidebar"] [data-testid="stImage"] {
    background: #ffffff !important;
    border-radius: 12px !important;
    padding: 10px 12px !important;
    margin-bottom: 6px !important;
}
section[data-testid="stSidebar"] [data-testid="stImage"] img {
    border-radius: 8px !important;
}

/* ── DIVIDERS ────────────────────────────────────────────────────────────────── */
hr { border-color: #e8f5e9 !important; }

/* ── DATAFRAME ───────────────────────────────────────────────────────────────── */
[data-testid="stDataFrame"] { border: 1px solid #e0f0e5 !important; border-radius: 8px !important; }

/* ── EXPANDER ────────────────────────────────────────────────────────────────── */
[data-testid="stExpander"] {
    border: 1px solid #e0f0e5 !important; border-radius: 8px !important;
    background: #f9fdf9 !important;
}
[data-testid="stExpander"] summary { color: #2d7a3a !important; font-weight: 600 !important; }

/* ── INFO / ALERT BOX ────────────────────────────────────────────────────────── */
[data-testid="stAlert"] { border-radius: 8px !important; border-left: 4px solid #2d7a3a !important; }

/* ── BUTTONS ─────────────────────────────────────────────────────────────────── */
[data-testid="stDownloadButton"] button {
    background-color: #2d7a3a !important; color: white !important;
    border-radius: 8px !important; border: none !important;
    font-weight: 600 !important;
}
[data-testid="stDownloadButton"] button:hover { background-color: #1a5c2a !important; }

/* ── FOOTER ──────────────────────────────────────────────────────────────────── */
.footer-bar {
    text-align: center; padding: 10px; font-size: 0.75rem;
    color: #8aab8e; border-top: 1px solid #e8f5e9; margin-top: 12px;
}

/* ── MOBILE ───────────────────────────────────────────────────────────────────── */
@media (max-width: 768px) {
    .block-container { padding-left: 0.8rem !important; padding-right: 0.8rem !important; }
    /* Boost Plotly chart text size */
    .js-plotly-plot .plotly text { font-size: 11px !important; }
    .js-plotly-plot .plotly .xtick text,
    .js-plotly-plot .plotly .ytick text { font-size: 10px !important; }
    /* Metric cards stack with comfortable padding */
    [data-testid="stMetric"] { padding: 12px 14px !important; }
    [data-testid="stMetricValue"] { font-size: 1.15rem !important; }
}
</style>
""", unsafe_allow_html=True)


# ── DATA LOADERS ────────────────────────────────────────────────────────────────

@st.cache_data
def load_expenses():
    wb = openpyxl.load_workbook(DATA_PATH)
    ws = wb["EXPENSES"]
    # New layout (row 4 = header, data from row 5):
    # col[0]=S/N, col[1]=DATE, col[2]=CATEGORY, col[3]=DESCRIPTION,
    # col[4]=PROCURED BY, col[5]=QTY, col[6]=SIZE/SPEC, col[7]=UNIT RATE, col[8]=AMOUNT
    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not any(v is not None for v in row):
            continue
        row = list(row) + [None] * 10
        date_val, category, desc, procured, qty, size, unit_rate, amount = row[1:9]
        if not isinstance(amount, (int, float)) or amount <= 0:
            continue
        if not isinstance(date_val, datetime):
            continue
        cat = str(category or "OTHER").strip().upper()
        if cat == "PIGS":
            cat = "PIG"  # merge stray category
        rows.append({
            "date": pd.Timestamp(date_val),
            "stock": cat,
            "expense_type": str(desc or "OTHER").strip(),
            "procured_by": str(procured or "UNKNOWN").strip(),
            "qty": str(qty).strip() if qty is not None else "",
            "size": str(size).strip() if size else "",
            "amount": float(amount)
        })
    df = pd.DataFrame(rows)
    df["month"] = df["date"].dt.to_period("M").astype(str)
    df["month_label"] = df["date"].dt.strftime("%b %Y")
    stock_map = {
        "PIG": "Pig", "FISH": "Fish", "OTHER": "Operations",
        "PEN PROJECT": "Pen Construction", "SITES PROJECT": "Site Works",
    }
    df["category"] = df["stock"].map(stock_map).fillna("Other")
    return df


@st.cache_data
def load_fish():
    wb = openpyxl.load_workbook(DATA_PATH)
    ws = wb["FISH RECORDS"]
    all_rows = list(ws.iter_rows(values_only=True))

    def to_ts(val):
        if isinstance(val, datetime):
            return pd.Timestamp(val)
        if isinstance(val, str):
            try:
                return pd.to_datetime(val)
            except Exception:
                pass
        return None

    # TABLE 1 — SMOKED FISH (header row 6, data from row 7, idx 6+)
    # col[0]=S/N, col[1]=DATE, col[2]=SPEC, col[3]=STOCK TYPE, col[4]=QTY, col[5]=# PACKS
    smoked = []
    for row in all_rows[6:]:
        row = list(row) + [None] * 8
        ts = to_ts(row[1])
        qty = row[4]
        if ts is None or not isinstance(qty, (int, float)) or qty <= 0:
            continue
        if ts.year < 2025 or ts.year > 2027:
            continue
        packs = row[5] if isinstance(row[5], (int, float)) else None
        smoked.append({"date": ts, "qty": float(qty), "packs": packs})

    smoked_df = pd.DataFrame(smoked)
    if not smoked_df.empty:
        smoked_df = smoked_df.drop_duplicates(subset=["date", "qty"]).sort_values("date").reset_index(drop=True)
        smoked_df["month_label"] = smoked_df["date"].dt.strftime("%b %Y")
        smoked_df["month"] = smoked_df["date"].dt.to_period("M").astype(str)

    # TABLE 2 — FINGERLINGS (header row 45, data from row 46, idx 45+)
    # col[0]=S/N, col[1]=POND, col[2]=QTY, col[3]=ARRIVAL DATE,
    # col[4]=BREED, col[5]=SIZE/STAGE, col[6]=TREATMENT, col[7]=TRANSPORT COST, col[8]=LOCATION
    fingerlings = []
    for i, row in enumerate(all_rows[45:]):
        row = list(row) + [None] * 12
        qty = row[2]
        pond = row[1]
        if not isinstance(qty, (int, float)) or qty <= 0:
            continue
        arrival_ts = to_ts(row[3])
        transport_raw = row[7]
        transport_val = float(transport_raw) if isinstance(transport_raw, (int, float)) else 0.0
        fingerlings.append({
            "pond": f"Pond {str(pond).strip()}" if pond else f"Batch {i+1}",
            "qty": int(qty),
            "arrival": arrival_ts,
            "breed": str(row[4] or "Unknown").strip(),
            "size": str(row[5] or "Fingerlings").strip(),
            "transport": transport_val,
            "location": str(row[8] or "Unknown").strip(),
        })

    fingerlings_df = pd.DataFrame(fingerlings) if fingerlings else pd.DataFrame(
        columns=["pond", "qty", "arrival", "breed", "size", "transport", "location"])
    return smoked_df, fingerlings_df


@st.cache_data
def load_pigs():
    wb = openpyxl.load_workbook(DATA_PATH)
    ws = wb["PIG RECORDS"]
    all_rows = list(ws.iter_rows(values_only=True))

    def to_ts(val):
        if isinstance(val, datetime):
            return pd.Timestamp(val)
        if isinstance(val, str):
            try:
                return pd.to_datetime(val)
            except Exception:
                pass
        return None

    # ── TABLE 1 — INVENTORY (rows 6–10, idx 5–9) ────────────────────────────────
    # col[0]=S/N, col[1]=STOCK TYPE, col[2]=BREED, col[3]=MALES, col[4]=FEMALES, col[5]=TOTAL
    inventory = []
    for row in all_rows[5:10]:
        row = list(row) + [None] * 8
        sn = row[0]
        if not isinstance(sn, (int, float)):
            continue
        stock = str(row[1] or "").strip().title()
        breed = str(row[2] or "").strip()
        males   = row[3]; females = row[4]; total = row[5]
        inventory.append({
            "type":    stock,
            "breed":   breed,
            "males":   int(males)   if isinstance(males,   (int, float)) else 0,
            "females": int(females) if isinstance(females, (int, float)) else 0,
            "total":   int(total)   if isinstance(total,   (int, float)) else 0,
        })

    # ── TABLE 2 — PIG PROFILES (rows 18–27, idx 17–26) ──────────────────────────
    # col[0]=S/N, col[1]=NAME, col[2]=STOCK, col[3]=BREED, col[4]=DATE OF BIRTH,
    # col[5]=CROSS BREED/PARTNER, col[6]=ARRIVAL DATE, col[7]=GENDER,
    # col[8]=PURCHASE PRICE, col[9]=TAG NO
    profiles = []
    for row in all_rows[17:28]:
        row = list(row) + [None] * 12
        sn = row[0]
        if not isinstance(sn, (int, float)):
            continue
        name = str(row[1] or "").strip()
        if not name:
            continue
        breed   = str(row[3] or "").strip()
        dob     = to_ts(row[4])
        arrival = to_ts(row[6])
        gender  = str(row[7] or "").strip()
        price   = float(row[8]) if isinstance(row[8], (int, float)) else None
        tag_raw = row[9]
        tag = str(int(tag_raw)) if isinstance(tag_raw, (int, float)) else str(tag_raw or "").strip() or None
        profiles.append({
            "name": name, "breed": breed, "dob": dob, "arrival": arrival,
            "gender": gender, "price": price, "tag": tag
        })

    profiles_df = pd.DataFrame(profiles) if profiles else pd.DataFrame(
        columns=["name", "breed", "dob", "arrival", "gender", "price", "tag"])
    if not profiles_df.empty:
        profiles_df["dob"]     = pd.to_datetime(profiles_df["dob"],     errors="coerce")
        profiles_df["arrival"] = pd.to_datetime(profiles_df["arrival"], errors="coerce")
        now = pd.Timestamp.now()
        profiles_df["age_months"] = profiles_df["dob"].apply(
            lambda x: int((now - x).days / 30.44) if pd.notna(x) else None)

    # ── EVENTS FROM NAMED TABLES ────────────────────────────────────────────────
    events = []

    # TABLE 3 — FARROWING / BIRTH (rows 32+, idx 31+)
    # col[0]=S/N, col[1]=DATE, col[2]=DAY, col[3]=MOTHER,
    # col[4]=MALE PIGLETS, col[5]=FEMALE PIGLETS, col[6]=TOTAL BORN
    for row in all_rows[31:38]:
        row = list(row) + [None] * 8
        sn = row[0]
        if not isinstance(sn, (int, float)):
            continue
        ts = to_ts(row[1])
        if ts is None:
            continue
        mother  = str(row[3] or "Sow").strip()
        males   = int(row[4]) if isinstance(row[4], (int, float)) else None
        females = int(row[5]) if isinstance(row[5], (int, float)) else None
        total   = (males or 0) + (females or 0)
        desc = f"{mother} farrowed"
        if males is not None or females is not None:
            desc += f" — {males or '?'} males, {females or '?'} females ({total} total)"
        events.append({"date": ts, "type": "Birth", "icon": "🐣",
                        "description": desc, "detail": ""})

    # TABLE 4 — VACCINATION & TREATMENT (rows 38–43, idx 37–42)
    # col[0]=S/N, col[1]=DATE, col[2]=ANIMAL TYPE, col[3]=QUANTITY,
    # col[4]=DRUG NAME, col[5]=DRUG DETAIL, col[6]=AMOUNT, col[7]=NOTES
    for row in all_rows[37:44]:
        row = list(row) + [None] * 10
        sn = row[0]
        if not isinstance(sn, (int, float)):
            continue
        ts = to_ts(row[1])
        if ts is None:
            continue
        animal = str(row[2] or "").strip()
        qty    = row[3]
        drug   = str(row[4] or "").strip()
        detail = str(row[5] or "").strip()
        amount = row[6]
        notes  = str(row[7] or "").strip()
        if not drug or drug == "-":
            drug = notes or "Treatment"
        desc = drug + (f" — {animal}" if animal else "")
        if qty and isinstance(qty, (int, float)):
            desc += f" ({int(qty)} animals)"
        detail_str = detail
        if isinstance(amount, (int, float)):
            detail_str += f" | Cost: ₦{amount:,.0f}"
        if notes and notes not in detail_str:
            detail_str = (detail_str + " | " + notes).strip(" | ")
        events.append({"date": ts, "type": "Vaccination", "icon": "💉",
                        "description": desc, "detail": detail_str.strip()})

    # TABLE 5 — MEDICATION (rows 49–54, idx 48–53)
    # col[0]=S/N, col[1]=DATE, col[2]=DRUG/MEDICINE, col[3]=TREATMENT TYPE,
    # col[4]=DURATION, col[5]=DOSAGE, col[6]=PIG(S), col[7]=ADMINISTERED BY
    for row in all_rows[48:55]:
        row = list(row) + [None] * 10
        sn = row[0]
        if not isinstance(sn, (int, float)):
            continue
        ts = to_ts(row[1])
        if ts is None:
            continue
        drug      = str(row[2] or "").strip()
        treatment = str(row[3] or "").strip()
        duration  = str(row[4] or "").strip()
        dosage    = str(row[5] or "").strip()
        pig       = str(row[6] or "").strip()
        admin     = str(row[7] or "").strip()
        if not drug:
            continue
        desc = drug + (f" → {pig}" if pig else "")
        parts = [p for p in [treatment, f"Dose: {dosage}" if dosage else "", f"Duration: {duration}" if duration else "", f"By: {admin}" if admin else ""] if p]
        events.append({"date": ts, "type": "Medication", "icon": "💊",
                        "description": desc, "detail": " | ".join(parts)})

    # TABLE 6 — HEAT / MATING (rows 58+, idx 57+)
    # col[0]=S/N, col[1]=DATE, col[2]=PIG NAME, col[3]=MATING DURATION,
    # col[4]=CROSSED WITH, col[5]=TAG NO
    for row in all_rows[57:65]:
        row = list(row) + [None] * 8
        sn = row[0]
        if not isinstance(sn, (int, float)):
            continue
        ts = to_ts(row[1])
        if ts is None:
            continue
        pig_name     = str(row[2] or "").strip()
        duration     = str(row[3] or "").strip()
        crossed_with = str(row[4] or "").strip()
        tag_raw      = row[5]
        tag = str(int(tag_raw)) if isinstance(tag_raw, (int, float)) else str(tag_raw or "").strip()
        if not pig_name:
            continue
        desc = f"{pig_name.upper()} mated" + (f" × {crossed_with}" if crossed_with else "")
        if tag:
            desc += f" (Tag {tag})"
        expected = ts + pd.Timedelta(days=114)
        detail_parts = []
        if duration:
            detail_parts.append(f"Period: {duration}")
        detail_parts.append(f"Expected farrowing: ~{expected.strftime('%d %b %Y')} (114 days)")
        events.append({"date": ts, "type": "Mating", "icon": "❤️",
                        "description": desc, "detail": " | ".join(detail_parts)})

    events_df = pd.DataFrame(events) if events else pd.DataFrame(
        columns=["date", "type", "icon", "description", "detail"])
    if not events_df.empty:
        events_df = (events_df
                     .drop_duplicates(subset=["date", "type", "description"])
                     .sort_values("date")
                     .reset_index(drop=True))

    return pd.DataFrame(inventory), profiles_df, events_df


@st.cache_data
def load_sales():
    wb = openpyxl.load_workbook(DATA_PATH)
    ws = wb["SALES"]
    all_rows = list(ws.iter_rows(values_only=True))
    # col[0]=S/N, col[1]=DATE, col[2]=LOCATION/REP, col[3]=QTY (PACKS)
    sales = []
    for row in all_rows[3:]:
        row = list(row) + [None] * 6
        qty = row[3]
        if not isinstance(qty, (int, float)) or qty <= 0:
            continue
        ts = pd.Timestamp(row[1]) if isinstance(row[1], datetime) else None
        location = str(row[2] or "Unknown").strip()
        sales.append({"date": ts, "location": location, "packs": float(qty)})
    sales_df = pd.DataFrame(sales) if sales else pd.DataFrame(
        columns=["date", "location", "packs"])
    if not sales_df.empty:
        sales_df["month_label"] = sales_df["date"].apply(
            lambda x: x.strftime("%b %Y") if pd.notna(x) else "Unknown")
    return sales_df


# Load everything
df_exp = load_expenses()
smoked_df, fingerlings_df = load_fish()
inventory_df, profiles_df, events_df = load_pigs()
sales_df = load_sales()

COLORS = {
    "Pig": "#e07b39", "Fish": "#2196f3", "Operations": "#7b68ee",
    "Pen Construction": "#4caf50", "Site Works": "#ff9800",
    "General Stock": "#9e9e9e", "Other": "#bdbdbd"
}

# ── SIDEBAR ─────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.image(LOGO_PATH, use_container_width=True)
    st.markdown("### Navigation")
    page = st.radio(
        "", ["Farm Overview", "Expenses", "Livestock", "Fish Production", "Sales"],
        label_visibility="collapsed"
    )
    st.markdown("---")
    st.markdown(f"**Last updated:** `{datetime.now().strftime('%d %b %Y')}`")
    st.markdown(f"**Expense records:** `{len(df_exp)}`")
    st.markdown(f"**Total animals:** `{inventory_df['total'].sum()}`")
    st.markdown(f"**Fish stocked:** `{fingerlings_df['qty'].sum():,}`")
    total_smoked_sidebar = smoked_df["qty"].sum() if not smoked_df.empty else 0
    st.markdown(f"**Fish output:** `{total_smoked_sidebar:,.0f}`")

# ── Mobile: auto-collapse sidebar when nav item is clicked ──────────────────
components.html("""
<script>
(function() {
    function setup() {
        var p = window.parent;
        var doc = p.document;
        var labels = doc.querySelectorAll('[data-testid="stSidebar"] label');
        labels.forEach(function(lbl) {
            if (lbl._mobileCollapseSet) return;
            lbl._mobileCollapseSet = true;
            lbl.addEventListener('click', function() {
                if (p.innerWidth <= 768) {
                    setTimeout(function() {
                        var btn = doc.querySelector('[data-testid="stSidebarCollapseButton"] button');
                        if (btn) btn.click();
                    }, 350);
                }
            });
        });
    }
    setup();
    setTimeout(setup, 800);
})();
</script>
""", height=0)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: FARM OVERVIEW
# ══════════════════════════════════════════════════════════════════════════════
if page == "Farm Overview":
    st.markdown("# FarmHub Agro Services — Farm Overview")
    st.markdown("*High-level view of all operations: livestock, fish, and spend*")
    st.markdown("---")

    this_month = datetime.now().strftime("%b %Y")
    total_spend = df_exp["amount"].sum()
    this_month_spend = df_exp[df_exp["month_label"] == this_month]["amount"].sum()
    total_smoked = smoked_df["qty"].sum() if not smoked_df.empty else 0
    this_month_smoked = smoked_df[smoked_df["month_label"] == this_month]["qty"].sum() if not smoked_df.empty else 0
    total_pigs = inventory_df["total"].sum()
    total_fish_stocked = fingerlings_df["qty"].sum()

    c1, c2, c3 = st.columns(3)
    with c1: st.metric("Total Farm Spend", f"₦{total_spend:,.0f}")
    with c2: st.metric("This Month Spend", f"₦{this_month_spend:,.0f}")
    with c3: st.metric("Total Animals", f"{total_pigs} pigs")
    c4, c5, c6 = st.columns(3)
    with c4: st.metric("Fish Stocked", f"{total_fish_stocked:,}")
    with c5: st.metric("Total Fish Output", f"{total_smoked:,.0f}")
    with c6: st.metric("Fish Output (Month)", f"{this_month_smoked:,.0f}")

    st.markdown("---")

    col1, col2 = st.columns(2)
    with col1:
        st.markdown('<div class="section-header">Monthly Farm Spend by Category</div>', unsafe_allow_html=True)
        monthly_exp = df_exp.groupby(["month_label", "category"])["amount"].sum().reset_index()
        monthly_exp["sort_key"] = pd.to_datetime(monthly_exp["month_label"], format="%b %Y")
        monthly_exp = monthly_exp.sort_values("sort_key")
        fig = px.bar(monthly_exp, x="month_label", y="amount", color="category",
                     color_discrete_map=COLORS, barmode="stack",
                     labels={"amount": "Amount (₦)", "month_label": "Month", "category": "Category"})
        fig.update_layout(height=300, plot_bgcolor="white", paper_bgcolor="white",
                          margin=dict(l=10, r=10, t=10, b=10),
                          yaxis=dict(tickformat=",.0f", gridcolor="#f0f0f0", automargin=True),
                          legend=dict(orientation="h", yanchor="bottom", y=1.02, x=0))
        fig.update_traces(marker_line_width=0)
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.markdown('<div class="section-header">Monthly Fish Output (Smoked)</div>', unsafe_allow_html=True)
        if not smoked_df.empty:
            monthly_fish = smoked_df.groupby("month_label")["qty"].sum().reset_index()
            monthly_fish["sort_key"] = pd.to_datetime(monthly_fish["month_label"], format="%b %Y")
            monthly_fish = monthly_fish.sort_values("sort_key")
            fig = px.bar(monthly_fish, x="month_label", y="qty",
                         color_discrete_sequence=["#2196f3"],
                         labels={"qty": "Smoked Fish Output", "month_label": "Month"})
            fig.update_layout(height=300, plot_bgcolor="white", paper_bgcolor="white",
                              margin=dict(l=10, r=10, t=10, b=10),
                              yaxis=dict(gridcolor="#f0f0f0", automargin=True), xaxis=dict(gridcolor="#f0f0f0", automargin=True))
            fig.update_traces(marker_line_width=0)
            st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")
    col1, col2 = st.columns(2)

    with col1:
        st.markdown('<div class="section-header">Livestock Inventory by Breed</div>', unsafe_allow_html=True)
        fig = px.bar(inventory_df, x="breed", y=["males", "females"],
                     barmode="group", color_discrete_sequence=["#2196f3", "#f48fb1"],
                     labels={"value": "Count", "breed": "Breed", "variable": "Gender"})
        fig.update_layout(height=260, plot_bgcolor="white", paper_bgcolor="white",
                          margin=dict(l=10, r=10, t=10, b=10),
                          yaxis=dict(gridcolor="#f0f0f0"), legend=dict(title=""))
        fig.update_traces(marker_line_width=0)
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.markdown('<div class="section-header">Key Farm Events</div>', unsafe_allow_html=True)

        # Build combined event list: pig events + fish stocking events
        all_events = events_df.copy() if not events_df.empty else pd.DataFrame(
            columns=["date", "type", "icon", "description", "detail"])

        # Add fish stocking events
        fish_events = []
        for _, fr in fingerlings_df.iterrows():
            if fr["arrival"] is not None:
                arr = fr["arrival"]
                transport_str = f" | Transport: ₦{fr['transport']:,.0f}" if fr["transport"] > 0 else ""
                fish_events.append({
                    "date": arr, "type": "Restocking", "icon": "🐠",
                    "description": f"{fr['qty']:,} {fr['breed']} stocked — {fr['pond']}",
                    "detail": f"Source: {fr['location']}{transport_str}",
                })
        if fish_events:
            all_events = pd.concat([all_events, pd.DataFrame(fish_events)], ignore_index=True)

        if not all_events.empty:
            all_events = all_events.sort_values("date").reset_index(drop=True)
            # Split past vs upcoming
            now_ts = pd.Timestamp.now()
            past   = all_events[all_events["date"] <= now_ts].tail(6)
            future = all_events[all_events["date"] >  now_ts]

            if not past.empty:
                for _, ev in past.iloc[::-1].iterrows():  # most recent first
                    date_str = ev["date"].strftime("%d %b %Y")
                    detail_html = f"<br><small style='color:#666'>{ev['detail']}</small>" if ev["detail"] else ""
                    st.markdown(
                        f'<div class="insight-card"><b>{date_str}</b> — {ev["description"]}{detail_html}</div>',
                        unsafe_allow_html=True)
            if not future.empty:
                st.markdown("**Upcoming**")
                for _, ev in future.iterrows():
                    date_str = ev["date"].strftime("%d %b %Y")
                    detail_html = f"<br><small style='color:#666'>{ev['detail']}</small>" if ev["detail"] else ""
                    st.markdown(
                        f'<div class="insight-card" style="border-left-color:#ff9800;"><b>{date_str}</b> — {ev["description"]}{detail_html}</div>',
                        unsafe_allow_html=True)
        else:
            st.info("No events found in the data.")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: EXPENSES
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Expenses":
    st.markdown("# Expense Analytics")
    st.markdown("*Full visibility into farm spend across all categories*")
    st.markdown("---")

    with st.expander("Filters", expanded=True):
        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            month_labels = sorted(df_exp["month_label"].unique(), key=lambda x: pd.to_datetime(x, format="%b %Y"))
            selected_months = st.multiselect("Month", options=month_labels, default=month_labels)
        with fc2:
            categories = sorted(df_exp["category"].unique())
            selected_cats = st.multiselect("Category", options=categories, default=categories)
        with fc3:
            procurers = sorted(df_exp["procured_by"].unique())
            selected_proc = st.multiselect("Procured By", options=procurers, default=procurers)

    filtered = df_exp[
        df_exp["month_label"].isin(selected_months) &
        df_exp["category"].isin(selected_cats) &
        df_exp["procured_by"].isin(selected_proc)
    ]

    total_spend = filtered["amount"].sum()
    num_tx = len(filtered)
    top_category = filtered.groupby("category")["amount"].sum().idxmax() if num_tx > 0 else "N/A"
    monthly_totals = filtered.groupby("month")["amount"].sum().sort_index()
    avg_monthly = monthly_totals.mean() if len(monthly_totals) > 0 else 0
    pig_spend = filtered[filtered["category"] == "Pig"]["amount"].sum()
    fish_spend = filtered[filtered["category"] == "Fish"]["amount"].sum()
    pig_pct = pig_spend / total_spend * 100 if total_spend > 0 else 0
    fish_pct = fish_spend / total_spend * 100 if total_spend > 0 else 0

    cur_month_spend = monthly_totals.iloc[-1] if len(monthly_totals) >= 1 else 0
    if len(monthly_totals) >= 2:
        prev = monthly_totals.iloc[-2]
        mom_diff = cur_month_spend - prev
        mom_sign = "+" if mom_diff >= 0 else "-"
        delta_label = f"{mom_sign}₦{abs(mom_diff):,.0f} vs prev month"
        avg_diff = cur_month_spend - avg_monthly
        avg_sign = "+" if avg_diff >= 0 else "-"
        avg_delta = f"{avg_sign}₦{abs(avg_diff):,.0f} vs avg"
    else:
        delta_label = avg_delta = None

    c1, c2, c3 = st.columns(3)
    with c1: st.metric("Total Spend", f"₦{total_spend:,.0f}",
                       delta=delta_label, delta_color="inverse",
                       help="Delta shows current month vs prior month. Green = spending decreased (good).")
    with c2: st.metric("Avg Monthly", f"₦{avg_monthly:,.0f}",
                       delta=avg_delta, delta_color="inverse",
                       help="Delta shows current month vs monthly average. Green = below average spend.")
    with c3: st.metric("Transactions", f"{num_tx}")
    c4, c5 = st.columns(2)
    with c4: st.metric("Top Category", top_category)
    with c5: st.metric("Pig / Fish Feed", f"{pig_pct:.0f}% / {fish_pct:.0f}%")

    st.markdown("---")

    # Insights
    if num_tx > 0:
        feed_kw = ["TOP FEEDS", "SKRETTING", "CORN BROWN", "CONR BROWN", "BSSM", "BSSERT", "C/B", "COPPENS"]
        feed_spend = filtered[filtered["expense_type"].str.upper().str.contains("|".join(feed_kw), na=False)]["amount"].sum()
        feed_pct = feed_spend / total_spend * 100 if total_spend > 0 else 0
        top_month_label = pd.Period(monthly_totals.idxmax(), freq="M").strftime("%b %Y")
        top_month_val = monthly_totals.max()
        top_agent = filtered.groupby("procured_by")["amount"].sum().idxmax()
        top_agent_val = filtered.groupby("procured_by")["amount"].sum().max()
        top_agent_pct = top_agent_val / total_spend * 100 if total_spend > 0 else 0

        ic1, ic2, ic3 = st.columns(3)
        with ic1:
            st.markdown(f'<div class="insight-card">Feed costs: <b>₦{feed_spend:,.0f}</b> ({feed_pct:.1f}% of total)</div>', unsafe_allow_html=True)
        with ic2:
            st.markdown(f'<div class="insight-card">Highest spend month: <b>{top_month_label}</b> — ₦{top_month_val:,.0f}</div>', unsafe_allow_html=True)
        with ic3:
            st.markdown(f'<div class="insight-card">Top procurer: <b>{top_agent.title()}</b> — ₦{top_agent_val:,.0f} ({top_agent_pct:.1f}%)</div>', unsafe_allow_html=True)
        st.markdown("---")

    col1, col2 = st.columns([2, 1])
    with col1:
        st.markdown('<div class="section-header">Monthly Spend Trend by Category</div>', unsafe_allow_html=True)
        monthly_cat = filtered.groupby(["month_label", "category"])["amount"].sum().reset_index()
        monthly_cat["sort_key"] = pd.to_datetime(monthly_cat["month_label"], format="%b %Y")
        monthly_cat = monthly_cat.sort_values("sort_key")
        fig = px.bar(monthly_cat, x="month_label", y="amount", color="category",
                     color_discrete_map=COLORS, barmode="stack",
                     labels={"amount": "Amount (₦)", "month_label": "Month"})
        fig.update_layout(height=320, plot_bgcolor="white", paper_bgcolor="white",
                          margin=dict(l=10, r=10, t=10, b=10),
                          yaxis=dict(tickformat=",.0f", gridcolor="#f0f0f0", automargin=True),
                          legend=dict(orientation="h", yanchor="bottom", y=1.02, x=0))
        fig.update_traces(marker_line_width=0)
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.markdown('<div class="section-header">Spend by Category</div>', unsafe_allow_html=True)
        cat_totals = filtered.groupby("category")["amount"].sum().reset_index()
        fig = px.pie(cat_totals, values="amount", names="category",
                     color="category", color_discrete_map=COLORS, hole=0.45)
        fig.update_layout(height=320, margin=dict(l=10, r=10, t=10, b=60),
                          paper_bgcolor="white",
                          legend=dict(orientation="h", yanchor="top", y=-0.08, xanchor="center", x=0.5))
        fig.update_traces(textposition="inside", textinfo="percent")
        st.plotly_chart(fig, use_container_width=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown('<div class="section-header">Feed Cost — Pig vs Fish (Monthly)</div>', unsafe_allow_html=True)
        feed_kw = ["TOP FEEDS", "SKRETTING", "CORN BROWN", "CONR BROWN", "BSSM", "BSSERT", "C/B", "COPPENS"]
        feed_df = filtered[filtered["expense_type"].str.upper().str.contains("|".join(feed_kw), na=False)].copy()
        feed_monthly = feed_df.groupby(["month_label", "category"])["amount"].sum().reset_index()
        feed_monthly["sort_key"] = pd.to_datetime(feed_monthly["month_label"], format="%b %Y")
        feed_monthly = feed_monthly.sort_values("sort_key")
        feed_monthly = feed_monthly[feed_monthly["category"].isin(["Pig", "Fish"])]
        fig = px.line(feed_monthly, x="month_label", y="amount", color="category",
                      color_discrete_map=COLORS, markers=True,
                      labels={"amount": "Feed Cost (₦)", "month_label": "Month"})
        fig.update_layout(height=280, plot_bgcolor="white", paper_bgcolor="white",
                          margin=dict(l=10, r=10, t=10, b=10),
                          yaxis=dict(tickformat=",.0f", gridcolor="#f0f0f0", automargin=True), legend=dict(title=""))
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.markdown('<div class="section-header">Spend by Procurement Agent</div>', unsafe_allow_html=True)
        proc_totals = filtered.groupby("procured_by")["amount"].sum().reset_index().sort_values("amount")
        fig = px.bar(proc_totals, x="amount", y="procured_by", orientation="h",
                     color_discrete_sequence=["#2d7a3a"],
                     labels={"amount": "Total (₦)", "procured_by": "Agent"})
        fig.update_layout(height=280, plot_bgcolor="white", paper_bgcolor="white",
                          margin=dict(l=10, r=10, t=10, b=10),
                          xaxis=dict(tickformat=",.0f", gridcolor="#f0f0f0"))
        fig.update_traces(marker_line_width=0)
        st.plotly_chart(fig, use_container_width=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown('<div class="section-header">Top 10 Expense Items</div>', unsafe_allow_html=True)
        top_items = filtered.groupby("expense_type")["amount"].sum().reset_index()
        top_items = top_items.sort_values("amount", ascending=False).head(10)
        top_items["expense_type"] = top_items["expense_type"].str.title().str[:32]
        fig = px.bar(top_items.sort_values("amount"), x="amount", y="expense_type", orientation="h",
                     color_discrete_sequence=["#e07b39"],
                     labels={"amount": "Total (₦)", "expense_type": "Item"})
        fig.update_layout(height=320, plot_bgcolor="white", paper_bgcolor="white",
                          margin=dict(l=10, r=10, t=10, b=10),
                          xaxis=dict(tickformat=",.0f", gridcolor="#f0f0f0"))
        fig.update_traces(marker_line_width=0)
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.markdown('<div class="section-header">Weekly Spend Pattern</div>', unsafe_allow_html=True)
        dow_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        dow_spend = (
            filtered.assign(day_of_week=filtered["date"].dt.day_name())
            .groupby("day_of_week")["amount"].sum()
            .reindex(dow_order).reset_index()
        )
        dow_spend.columns = ["day", "amount"]
        fig = px.bar(dow_spend, x="day", y="amount", color_discrete_sequence=["#2196f3"],
                     labels={"amount": "Total (₦)", "day": "Day"})
        fig.update_layout(height=320, plot_bgcolor="white", paper_bgcolor="white",
                          margin=dict(l=10, r=10, t=10, b=10),
                          yaxis=dict(tickformat=",.0f", gridcolor="#f0f0f0", automargin=True))
        fig.update_traces(marker_line_width=0)
        st.plotly_chart(fig, use_container_width=True)

    # Pen construction tracker
    st.markdown("---")
    st.markdown('<div class="section-header">Pen Construction — Cumulative Spend Tracker</div>', unsafe_allow_html=True)
    pen_df = df_exp[df_exp["category"] == "Pen Construction"].copy().sort_values("date")
    pen_df["cumulative"] = pen_df["amount"].cumsum()
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(go.Bar(x=pen_df["date"], y=pen_df["amount"], name="Per Transaction",
                         marker_color="#81c784", marker_line_width=0), secondary_y=False)
    fig.add_trace(go.Scatter(x=pen_df["date"], y=pen_df["cumulative"], name="Cumulative",
                             line=dict(color="#2d7a3a", width=2.5), mode="lines+markers"), secondary_y=True)
    fig.update_layout(height=280, plot_bgcolor="white", paper_bgcolor="white",
                      margin=dict(l=10, r=10, t=10, b=10),
                      legend=dict(orientation="h", yanchor="bottom", y=1.02))
    fig.update_yaxes(tickformat=",.0f", gridcolor="#f0f0f0")
    st.plotly_chart(fig, use_container_width=True)
    pen_total = pen_df["amount"].sum()
    st.info(f"Pen construction total: **₦{pen_total:,.0f}** | Transactions: **{len(pen_df)}** | "
            f"Period: **{pen_df['date'].min().strftime('%d %b %Y')} – {pen_df['date'].max().strftime('%d %b %Y')}**")

    # Export
    st.markdown("---")
    with st.expander("View & Export Transaction Data"):
        display_df = filtered[["date", "category", "expense_type", "procured_by", "qty", "size", "amount"]].copy()
        display_df["date"] = display_df["date"].dt.strftime("%d %b %Y")
        display_df["qty"] = display_df["qty"].astype(str)
        display_df["amount_fmt"] = display_df["amount"].apply(lambda x: f"₦{x:,.0f}")
        show_df = display_df.drop(columns=["amount"]).rename(columns={
            "date": "Date", "category": "Category", "expense_type": "Expense",
            "procured_by": "Procured By", "qty": "Qty", "size": "Size", "amount_fmt": "Amount"
        })
        st.dataframe(show_df, use_container_width=True, height=300)
        csv_out = filtered[["date", "category", "expense_type", "procured_by", "qty", "size", "amount"]].copy()
        csv_out["date"] = csv_out["date"].dt.strftime("%d %b %Y")
        csv_out["qty"] = csv_out["qty"].astype(str)
        st.download_button("Download CSV", data=csv_out.to_csv(index=False).encode("utf-8"),
                           file_name="farmhub_expenses.csv", mime="text/csv")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: LIVESTOCK
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Livestock":
    st.markdown("# Livestock Management")
    st.markdown("*Pig inventory, individual animal profiles, health & breeding records*")
    st.markdown("---")

    total_pigs = inventory_df["total"].sum()
    adults = inventory_df[inventory_df["type"] == "Pig"]["total"].sum()
    piglets = inventory_df[inventory_df["type"] == "Piglets"]["total"].sum()
    breeds = inventory_df["breed"].nunique()
    pig_spend = df_exp[df_exp["category"] == "Pig"]["amount"].sum()

    c1, c2, c3 = st.columns(3)
    with c1: st.metric("Total Animals", total_pigs)
    with c2: st.metric("Adult Pigs", adults)
    with c3: st.metric("Piglets", piglets)
    c4, c5 = st.columns(2)
    with c4: st.metric("Breeds", breeds)
    with c5: st.metric("Total Pig Spend", f"₦{pig_spend:,.0f}")

    st.markdown("---")

    col1, col2 = st.columns(2)
    with col1:
        st.markdown('<div class="section-header">Stock by Breed & Gender</div>', unsafe_allow_html=True)
        fig = px.bar(inventory_df, x="breed", y=["males", "females"],
                     barmode="group", color_discrete_sequence=["#2196f3", "#f48fb1"],
                     labels={"value": "Count", "breed": "Breed", "variable": "Gender"})
        fig.update_layout(height=260, plot_bgcolor="white", paper_bgcolor="white",
                          margin=dict(l=10, r=10, t=10, b=10),
                          yaxis=dict(gridcolor="#f0f0f0"), legend=dict(title=""))
        fig.update_traces(marker_line_width=0)
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.markdown('<div class="section-header">Stock Distribution</div>', unsafe_allow_html=True)
        fig = px.pie(inventory_df, values="total", names="breed", hole=0.4,
                     color_discrete_sequence=px.colors.qualitative.Set2)
        fig.update_layout(height=260, margin=dict(l=10, r=10, t=10, b=10), paper_bgcolor="white")
        fig.update_traces(textposition="inside", textinfo="label+percent")
        st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")
    st.markdown('<div class="section-header">Individual Animal Profiles</div>', unsafe_allow_html=True)
    cols = st.columns(4)
    for i, row in profiles_df.iterrows():
        with cols[i % 4]:
            gender_icon = "M" if "male" in str(row.get("gender") or "").lower() and "fe" not in str(row.get("gender") or "").lower() else "F"
            age_str = f"{row['age_months']} months" if pd.notna(row.get("age_months")) and row["age_months"] else "Unknown"
            price_str = f"₦{row['price']:,.0f}" if pd.notna(row.get("price")) and row["price"] else "—"
            tag_str = row["tag"] if pd.notna(row.get("tag")) and row["tag"] else "—"
            dob_str = row["dob"].strftime("%d %b %Y") if pd.notna(row.get("dob")) else "—"
            arrival_str = row["arrival"].strftime("%d %b %Y") if pd.notna(row.get("arrival")) else "—"
            st.markdown(f"""
            <div class="animal-card">
                <h4>{gender_icon} {row['name']}</h4>
                <b>Breed:</b> {row['breed']}<br>
                <b>DOB:</b> {dob_str}<br>
                <b>Age:</b> {age_str}<br>
                <b>Arrival:</b> {arrival_str}<br>
                <b>Tag No:</b> {tag_str}<br>
                <b>Purchase:</b> {price_str}
            </div>
            """, unsafe_allow_html=True)

    st.markdown("---")
    col1, col2 = st.columns(2)

    with col1:
        st.markdown('<div class="section-header">Health Events (Vaccinations & Medications)</div>', unsafe_allow_html=True)
        health_events = events_df[events_df["type"].isin(["Vaccination", "Medication"])] if not events_df.empty else pd.DataFrame()
        if not health_events.empty:
            for _, ev in health_events.sort_values("date", ascending=False).iterrows():
                date_str = ev["date"].strftime("%d %b %Y")
                detail_html = f"<br><small style='color:#666'>{ev['detail']}</small>" if ev["detail"] else ""
                st.markdown(
                    f'<div class="insight-card"><b>{date_str}</b> — {ev["description"]}{detail_html}</div>',
                    unsafe_allow_html=True)
        else:
            st.info("No health events found in the data.")

        st.markdown("**Pig Spend by Item**")
        pig_items = df_exp[df_exp["category"] == "Pig"].groupby("expense_type")["amount"].sum().reset_index()
        pig_items = pig_items.sort_values("amount", ascending=False).head(8)
        pig_items["expense_type"] = pig_items["expense_type"].str.title().str[:28]
        fig = px.bar(pig_items.sort_values("amount"), x="amount", y="expense_type", orientation="h",
                     color_discrete_sequence=["#e07b39"],
                     labels={"amount": "Total (₦)", "expense_type": "Item"})
        fig.update_layout(height=260, plot_bgcolor="white", paper_bgcolor="white",
                          margin=dict(l=10, r=10, t=10, b=10),
                          xaxis=dict(tickformat=",.0f", gridcolor="#f0f0f0"))
        fig.update_traces(marker_line_width=0)
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.markdown('<div class="section-header">Breeding & Birth Records</div>', unsafe_allow_html=True)
        breed_events = events_df[events_df["type"].isin(["Birth", "Mating"])] if not events_df.empty else pd.DataFrame()
        if not breed_events.empty:
            for _, ev in breed_events.sort_values("date", ascending=False).iterrows():
                date_str = ev["date"].strftime("%d %b %Y")
                detail_html = f"<br><small style='color:#666'>{ev['detail']}</small>" if ev["detail"] else ""
                st.markdown(
                    f'<div class="insight-card"><b>{date_str}</b> — {ev["description"]}{detail_html}</div>',
                    unsafe_allow_html=True)
        else:
            st.info("No breeding or birth records found in the data.")

        st.markdown("**Pig Spend Over Time**")
        pig_monthly = df_exp[df_exp["category"] == "Pig"].groupby("month_label")["amount"].sum().reset_index()
        pig_monthly["sort_key"] = pd.to_datetime(pig_monthly["month_label"], format="%b %Y")
        pig_monthly = pig_monthly.sort_values("sort_key")
        fig = px.line(pig_monthly, x="month_label", y="amount", markers=True,
                      color_discrete_sequence=["#e07b39"],
                      labels={"amount": "Spend (₦)", "month_label": "Month"})
        fig.update_layout(height=260, plot_bgcolor="white", paper_bgcolor="white",
                          margin=dict(l=10, r=10, t=10, b=10),
                          yaxis=dict(tickformat=",.0f", gridcolor="#f0f0f0", automargin=True))
        st.plotly_chart(fig, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: FISH PRODUCTION
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Fish Production":
    st.markdown("# Fish Production")
    st.markdown("*Fingerling stocking, smoked fish output, and production trends*")
    st.markdown("---")

    total_stocked = fingerlings_df["qty"].sum()
    total_smoked = smoked_df["qty"].sum() if not smoked_df.empty else 0
    fish_spend = df_exp[df_exp["category"] == "Fish"]["amount"].sum()
    avg_batch = smoked_df["qty"].mean() if not smoked_df.empty else 0
    batches = len(smoked_df)
    cost_per_kg = fish_spend / total_smoked if total_smoked > 0 else 0

    c1, c2, c3 = st.columns(3)
    with c1: st.metric("Fingerlings Stocked", f"{total_stocked:,}")
    with c2: st.metric("Total Output", f"{total_smoked:,.0f}")
    with c3: st.metric("Fish Feed Spend", f"₦{fish_spend:,.0f}")
    c4, c5, c6 = st.columns(3)
    with c4: st.metric("Cost per Unit Output", f"₦{cost_per_kg:,.0f}")
    with c5: st.metric("Avg Batch Size", f"{avg_batch:.0f}")
    with c6: st.metric("Smoking Batches", batches)

    st.markdown("---")

    col1, col2 = st.columns(2)
    with col1:
        st.markdown('<div class="section-header">Smoked Fish Output — Each Batch</div>', unsafe_allow_html=True)
        if not smoked_df.empty:
            sorted_smoked = smoked_df.sort_values("date")
            fig = px.bar(sorted_smoked, x="date", y="qty", color_discrete_sequence=["#2196f3"],
                         labels={"qty": "Smoked Fish Output", "date": "Date"})
            fig.update_layout(height=300, plot_bgcolor="white", paper_bgcolor="white",
                              margin=dict(l=10, r=10, t=10, b=10),
                              yaxis=dict(gridcolor="#f0f0f0", automargin=True), xaxis=dict(gridcolor="#f0f0f0", automargin=True))
            fig.update_traces(marker_line_width=0)
            st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.markdown('<div class="section-header">Monthly Output & Cumulative Trend</div>', unsafe_allow_html=True)
        if not smoked_df.empty:
            monthly_fish = smoked_df.groupby("month_label")["qty"].sum().reset_index()
            monthly_fish["sort_key"] = pd.to_datetime(monthly_fish["month_label"], format="%b %Y")
            monthly_fish = monthly_fish.sort_values("sort_key")
            monthly_fish["cumulative"] = monthly_fish["qty"].cumsum()
            fig = make_subplots(specs=[[{"secondary_y": True}]])
            fig.add_trace(go.Bar(x=monthly_fish["month_label"], y=monthly_fish["qty"],
                                 name="Monthly Output", marker_color="#64b5f6", marker_line_width=0), secondary_y=False)
            fig.add_trace(go.Scatter(x=monthly_fish["month_label"], y=monthly_fish["cumulative"],
                                     name="Cumulative", line=dict(color="#1565c0", width=2.5),
                                     mode="lines+markers"), secondary_y=True)
            fig.update_layout(height=300, plot_bgcolor="white", paper_bgcolor="white",
                              margin=dict(l=10, r=10, t=10, b=10),
                              legend=dict(orientation="h", yanchor="bottom", y=1.02))
            fig.update_yaxes(gridcolor="#f0f0f0")
            st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")
    col1, col2 = st.columns(2)

    with col1:
        st.markdown('<div class="section-header">Fingerling Stocking Records</div>', unsafe_allow_html=True)
        if fingerlings_df.empty:
            st.info("No fingerling records found.")
        else:
            for _, row in fingerlings_df.iterrows():
                transport_str = f"₦{row['transport']:,.0f}" if row["transport"] > 0 else "Nil"
                arrival_str = row["arrival"].strftime("%d %b %Y") if pd.notna(row.get("arrival")) and row["arrival"] else "Unknown"
                st.markdown(f"""
                <div class="insight-card">
                    <b>{row['pond']}</b> — {row['qty']:,} {row['breed']} ({row['size']})<br>
                    Arrival: <b>{arrival_str}</b> |
                    Source: <b>{row['location']}</b> |
                    Transport: <b>{transport_str}</b>
                </div>
                """, unsafe_allow_html=True)

        fig = px.bar(fingerlings_df, x="pond", y="qty",
                     color_discrete_sequence=["#26a69a"],
                     labels={"qty": "Fingerlings Stocked", "pond": "Pond"}, text="qty")
        fig.update_traces(texttemplate="%{text:,}", textposition="outside", marker_line_width=0)
        fig.update_layout(height=240, plot_bgcolor="white", paper_bgcolor="white",
                          margin=dict(l=10, r=10, t=10, b=10),
                          yaxis=dict(gridcolor="#f0f0f0"))
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.markdown('<div class="section-header">Fish Feed Spend Over Time</div>', unsafe_allow_html=True)
        fish_exp = df_exp[df_exp["category"] == "Fish"].groupby("month_label")["amount"].sum().reset_index()
        fish_exp["sort_key"] = pd.to_datetime(fish_exp["month_label"], format="%b %Y")
        fish_exp = fish_exp.sort_values("sort_key")
        fig = px.line(fish_exp, x="month_label", y="amount", markers=True,
                      color_discrete_sequence=["#2196f3"],
                      labels={"amount": "Feed Cost (₦)", "month_label": "Month"})
        fig.update_layout(height=200, plot_bgcolor="white", paper_bgcolor="white",
                          margin=dict(l=10, r=10, t=10, b=10),
                          yaxis=dict(tickformat=",.0f", gridcolor="#f0f0f0", automargin=True))
        st.plotly_chart(fig, use_container_width=True)

        st.markdown('<div class="section-header">Fish Feed Breakdown by Item</div>', unsafe_allow_html=True)
        fish_items = df_exp[df_exp["category"] == "Fish"].groupby("expense_type")["amount"].sum().reset_index()
        fish_items = fish_items.sort_values("amount", ascending=False)
        fig = px.pie(fish_items, values="amount", names="expense_type", hole=0.4,
                     color_discrete_sequence=px.colors.sequential.Blues_r)
        fig.update_layout(height=210, margin=dict(l=10, r=10, t=10, b=10), paper_bgcolor="white")
        fig.update_traces(textposition="inside", textinfo="label+percent")
        st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")
    with st.expander("Smoked Fish Batch Log"):
        display = smoked_df[["date", "month_label", "qty"]].copy()
        display["date"] = display["date"].dt.strftime("%d %b %Y")
        display = display.rename(columns={"date": "Date", "qty": "Qty", "month_label": "Month"})
        st.dataframe(display[["Date", "Month", "Qty"]], use_container_width=True, height=300)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE: SALES
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Sales":
    st.markdown("# Sales & Dispatch")
    st.markdown("*Fish dispatch log — packs sent to buyers and locations*")
    st.markdown("---")

    total_packs = sales_df["packs"].sum() if not sales_df.empty else 0
    total_locations = sales_df["location"].nunique() if not sales_df.empty else 0
    avg_dispatch = sales_df["packs"].mean() if not sales_df.empty else 0
    last_date = sales_df["date"].dropna().max() if not sales_df.empty else None

    c1, c2 = st.columns(2)
    with c1: st.metric("Total Packs Dispatched", f"{int(total_packs):,}")
    with c2: st.metric("Dispatch Locations", total_locations)
    c3, c4 = st.columns(2)
    with c3: st.metric("Avg per Dispatch", f"{avg_dispatch:.0f} packs")
    with c4: st.metric("Last Dispatch", last_date.strftime("%d %b %Y") if last_date else "—")

    st.markdown("---")

    if not sales_df.empty:
        col1, col2 = st.columns(2)
        with col1:
            st.markdown('<div class="section-header">Dispatches by Location</div>', unsafe_allow_html=True)
            loc_totals = sales_df.groupby("location")["packs"].sum().reset_index().sort_values("packs", ascending=False)
            fig = px.bar(loc_totals, x="location", y="packs",
                         color_discrete_sequence=["#2d7a3a"],
                         labels={"packs": "Packs", "location": "Location"},
                         text="packs")
            fig.update_traces(texttemplate="%{text:.0f}", textposition="outside", marker_line_width=0)
            fig.update_layout(height=300, plot_bgcolor="white", paper_bgcolor="white",
                              margin=dict(l=10, r=10, t=10, b=10),
                              yaxis=dict(gridcolor="#f0f0f0"))
            st.plotly_chart(fig, use_container_width=True)

        with col2:
            st.markdown('<div class="section-header">Location Share</div>', unsafe_allow_html=True)
            fig = px.pie(loc_totals, values="packs", names="location", hole=0.4,
                         color_discrete_sequence=px.colors.qualitative.Set2)
            fig.update_layout(height=300, margin=dict(l=10, r=10, t=10, b=10), paper_bgcolor="white")
            fig.update_traces(textposition="inside", textinfo="label+percent")
            st.plotly_chart(fig, use_container_width=True)

        st.markdown("---")
        st.markdown('<div class="section-header">Dispatch Log</div>', unsafe_allow_html=True)
        for _, row in sales_df.iterrows():
            date_str = row["date"].strftime("%d %b %Y") if pd.notna(row.get("date")) else "Date unknown"
            st.markdown(
                f'<div class="insight-card"><b>{date_str}</b> — {int(row["packs"])} packs → <b>{row["location"]}</b></div>',
                unsafe_allow_html=True)
    else:
        st.info("No sales records found.")


# ── FOOTER ──────────────────────────────────────────────────────────────────────
st.markdown(
    '<div class="footer-bar">FarmHub Agro Services &nbsp;|&nbsp; Built by Folarin Osuolale &nbsp;|&nbsp; Data Science & Analytics</div>',
    unsafe_allow_html=True
)
